import os
import sys
import shutil
import math
from datetime import datetime, timedelta, timezone
import threading
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, session, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, text
from sqlalchemy.orm import joinedload
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import FlaskForm, CSRFProtect
from flask_wtf.csrf import CSRFError, generate_csrf
from wtforms import StringField, PasswordField, SubmitField, DecimalField, IntegerField, SelectField, TextAreaField, HiddenField
from wtforms.validators import DataRequired, Length, NumberRange, ValidationError, Optional
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime as dt, date
import json
import io
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from pathlib import Path

# Licensing (optional in older deployments)
try:
    from app.license_manager import (
        get_license_status,
        try_activate_online,
        get_device_fingerprint,
        get_license_key,
        save_license_key,
        allows_telegram,
    )
    LICENSE_AVAILABLE = True
except Exception:
    LICENSE_AVAILABLE = False

# Scheduler (optional)
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    SCHEDULER_AVAILABLE = True
except Exception:
    BackgroundScheduler = None  # type: ignore[assignment]
    CronTrigger = None  # type: ignore[assignment]
    SCHEDULER_AVAILABLE = False
    print('[WARNING] APScheduler not available. Scheduler features are disabled.')

# Telegram Bot Integration
# In Docker builds, the module may live as app.telegram_bot (package) instead of telegram_bot (top-level).
try:
    from app.telegram_bot import initialize_telegram_bot, get_telegram_bot
    TELEGRAM_AVAILABLE = True
except Exception:
    try:
        from telegram_bot import initialize_telegram_bot, get_telegram_bot
        TELEGRAM_AVAILABLE = True
    except Exception:
        TELEGRAM_AVAILABLE = False
        print('[WARNING] Telegram bot module not available. Telegram monitoring is disabled.')

_telegram_start_lock = threading.Lock()
_telegram_started = False


def _start_telegram_bot_if_configured(app: Flask) -> None:
    global _telegram_started
    if _telegram_started or not TELEGRAM_AVAILABLE:
        return

    if LICENSE_AVAILABLE and not allows_telegram():
        return

    werkzeug_run_main = os.environ.get('WERKZEUG_RUN_MAIN')
    if werkzeug_run_main is not None and werkzeug_run_main.lower() != 'true':
        return

    bot_token = (os.environ.get('TELEGRAM_BOT_TOKEN', '') or '').strip()
    admin_chat_ids_str = (os.environ.get('TELEGRAM_ADMIN_CHAT_IDS', '') or '').strip()
    if not bot_token or not admin_chat_ids_str:
        return

    admin_chat_ids = [cid.strip() for cid in admin_chat_ids_str.split(',') if cid.strip()]
    if not admin_chat_ids:
        return

    with _telegram_start_lock:
        if _telegram_started:
            return

        try:
            telegram_bot = initialize_telegram_bot(bot_token, admin_chat_ids, app)
            if telegram_bot:
                telegram_bot.start_bot_async()
                _telegram_started = True
                print('✅ Telegram Bot started (auto)')
            else:
                print('❌ Telegram Bot init failed (auto)')
        except Exception as e:
            print(f'❌ Telegram Bot error (auto): {e}')

print('=' * 60)
print('KASIR TOKO SEMBAKO - WITH SIMPLE BACKUP SYSTEM')
print('=' * 60)

# ==================== TIMEZONE HELPER ====================

def get_local_now():
    """Get current time in local timezone"""
    # Gunakan waktu lokal sistem (otomatis detect timezone dari OS)
    return datetime.now()

def get_local_timezone_name():
    """Get local timezone name (WIB/WITA/WIT/etc)"""
    import time
    
    # Get offset dari UTC dalam detik
    if time.daylight:
        offset_sec = -time.altzone
    else:
        offset_sec = -time.timezone
    
    # Convert ke jam
    offset_hours = offset_sec // 3600
    
    # Map ke nama timezone Indonesia
    timezone_map = {
        7: 'WIB',   # UTC+7 - Waktu Indonesia Barat
        8: 'WITA',  # UTC+8 - Waktu Indonesia Tengah
        9: 'WIT',   # UTC+9 - Waktu Indonesia Timur
    }
    
    # Return nama timezone atau offset format
    if offset_hours in timezone_map:
        return timezone_map[offset_hours]
    else:
        # Untuk timezone lain, tampilkan UTC offset
        sign = '+' if offset_hours >= 0 else '-'
        return f'UTC{sign}{abs(offset_hours)}'

# Debug: Print current time saat startup
local_tz_name = get_local_timezone_name()
print(f'[TIME] Local System Time: {get_local_now().strftime("%Y-%m-%d %H:%M:%S")}')
print(f'[TIME] Detected Timezone: {local_tz_name}')
print('=' * 60)

# ==================== SIMPLE BACKUP ====================

def backup_database():
    """Backup database dengan error handling yang lebih baik"""
    # Get base directory (parent of app folder)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_file = os.path.join(base_dir, 'instance', 'kasir.db')
    backup_folder = os.path.join(base_dir, 'backups')
    
    # Validasi path
    if not os.path.exists(os.path.dirname(db_file)):
        print(f"[Backup] ✗ Folder instance tidak ditemukan")
        return False
    
    # Pastikan folder backup ada
    try:
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder, exist_ok=True)
            print(f"[Backup] Created folder: {backup_folder}")
    except PermissionError:
        print(f"[Backup] ✗ Tidak ada izin membuat folder backup")
        return False
    
    if os.path.exists(db_file):
        try:
            # Cek ukuran database
            db_size = os.path.getsize(db_file)
            if db_size == 0:
                print("[Backup] ⚠️ Database kosong, backup dibatalkan")
                return False
            
            # Nama file backup
            timestamp = get_local_now().strftime('%Y%m%d_%H%M%S')
            backup_file = f'{backup_folder}/kasir_backup_{timestamp}.db'
            
            # Copy database dengan verifikasi
            shutil.copy2(db_file, backup_file)
            
            # Verifikasi backup
            if os.path.exists(backup_file):
                backup_size = os.path.getsize(backup_file)
                if backup_size == db_size:
                    print(f'[Backup] ✓ {backup_file}')
                    print(f'[Backup] Size: {backup_size/1024:.1f} KB')
                    
                    # Hapus backup lama (simpan 10 terbaru)
                    try:
                        backups = sorted([
                            f for f in os.listdir(backup_folder) 
                            if f.startswith('kasir_backup_') and f.endswith('.db')
                        ])
                        if len(backups) > 10:
                            for old_backup in backups[:-10]:
                                old_path = os.path.join(backup_folder, old_backup)
                                os.remove(old_path)
                                print(f'[Backup] Hapus backup lama: {old_backup}')
                    except Exception as e:
                        print(f'[Backup] ⚠️ Gagal hapus backup lama: {e}')
                    
                    return True
                else:
                    print(f'[Backup] ✗ Ukuran backup tidak sesuai')
                    os.remove(backup_file)  # Hapus backup yang rusak
                    return False
            else:
                print(f'[Backup] ✗ File backup tidak terbuat')
                return False
                
        except Exception as e:
            print(f'[Backup] ✗ Error: {e}')
            return False
    else:
        print(f'[Backup] ✗ Database not found: {db_file}')
        return False

# ==================== FLASK APP ====================

# Get base directory
base_app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
db_path = os.path.join(base_app_dir, 'instance', 'kasir.db')

app = Flask(__name__)
app.config['SECRET_KEY'] = 'rahasia-sangat-rahasia-123456'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path.replace(os.sep, "/")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'

# License enforcement toggle (opt-in; sale builds set LICENSE_ENFORCE=true)
app.config['LICENSE_ENFORCE'] = (os.environ.get('LICENSE_ENFORCE', 'false') or 'false').lower() == 'true'

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
csrf = CSRFProtect(app)


def _get_pagination_args(default_per_page: int = 50, max_per_page: int = 200) -> tuple[int, int]:
    page = request.args.get('page', 1, type=int) or 1
    per_page = request.args.get('per_page', default_per_page, type=int) or default_per_page
    page = max(1, page)
    per_page = max(1, min(max_per_page, per_page))
    return page, per_page


def _paginate_query(query, *, page: int, per_page: int, order_by=None):
    if order_by is not None:
        query = query.order_by(order_by)

    total = query.order_by(None).count()
    pages = max(1, math.ceil(total / per_page)) if per_page else 1
    if page > pages:
        page = pages
    items = query.limit(per_page).offset((page - 1) * per_page).all()
    return {
        'items': items,
        'page': page,
        'per_page': per_page,
        'total': total,
        'pages': pages,
        'has_prev': page > 1,
        'has_next': page < pages,
        'prev_num': page - 1,
        'next_num': page + 1,
    }


@app.before_request
def _enforce_license_before_request():
    if not app.config.get('LICENSE_ENFORCE', True):
        return None
    if not LICENSE_AVAILABLE:
        return None

    # Allow activation page and static
    path = request.path or ''
    if path.startswith('/static/') or path.startswith('/license'):
        return None

    status = get_license_status()
    if status.ok:
        return None

    # Redirect everything to activation page
    return redirect(url_for('license_activate'))


@app.route('/license', methods=['GET', 'POST'])
@csrf.exempt
def license_activate():
    if not LICENSE_AVAILABLE:
        return (
            'License system not available in this build. Please contact the vendor.',
            500,
        )

    message = ''
    if request.method == 'POST':
        key = (request.form.get('license_key', '') or '').strip()
        if not key:
            message = 'License key wajib diisi.'
        else:
            save_license_key(key)
            st = try_activate_online(timeout_s=10)
            if st.ok:
                message = '✅ Aktivasi berhasil.'
            else:
                message = f'❌ Aktivasi gagal: {st.reason or "unknown"}'

    status = get_license_status(refresh=True)
    return render_template(
        'license/activate.html',
        status=status,
        message=message,
        device_fingerprint=get_device_fingerprint(),
        license_key=get_license_key(),
    )

# ==================== SCHEDULER - Daily Saldo Archive ====================
scheduler = BackgroundScheduler() if SCHEDULER_AVAILABLE else None

def archive_daily_saldo():
    """Archive daily saldo setiap hari jam 22:30."""
    with app.app_context():
        try:
            from datetime import date as date_class
            today = date_class.today()
            
            # Check if already archived for today
            existing = DailySaldoArchive.query.filter_by(tanggal=today).first()
            if existing:
                return
            
            # Calculate saldo
            masuk = db.session.query(db.func.coalesce(db.func.sum(KasMutasi.jumlah), 0)).filter(KasMutasi.jenis == 'masuk').scalar() or 0
            keluar = db.session.query(db.func.coalesce(db.func.sum(KasMutasi.jumlah), 0)).filter(KasMutasi.jenis == 'keluar').scalar() or 0
            saldo_akhir = float(masuk) - float(keluar)
            
            # Archive
            archive = DailySaldoArchive(
                tanggal=today,
                saldo_awal=0,
                total_masuk=float(masuk),
                total_keluar=float(keluar),
                saldo_akhir=saldo_akhir
            )
            db.session.add(archive)
            db.session.commit()
            print(f'[SCHEDULER] Daily saldo archived for {today}: Rp {saldo_akhir:,.0f}')
        except Exception as e:
            print(f'[SCHEDULER ERROR] Failed to archive daily saldo: {e}')

def generate_daily_report():
    """Generate laporan per hari untuk hari ini dan simpan ke folder reports."""
    with app.app_context():
        try:
            from datetime import date as date_class
            import os
            
            today = date_class.today()
            today_str = today.strftime('%Y-%m-%d')
            
            # Create reports directory if not exists
            reports_dir = os.path.join(app.instance_path, 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            
            # Generate report using existing function
            wb = generate_laporan_hari(today_str, today_str)
            
            # Save to file
            filename = f'laporan_harian_{today_str}.xlsx'
            filepath = os.path.join(reports_dir, filename)
            wb.save(filepath)
            
            print(f'[SCHEDULER] Daily report generated: {filepath}')
        except Exception as e:
            print(f'[SCHEDULER ERROR] Failed to generate daily report: {e}')

def ensure_db_columns():
    """Ensure all required columns exist in database (auto-migration)."""
    try:
        with app.app_context():
            # Check if columns exist, add them if not
            from sqlalchemy import inspect, text
            
            inspector = inspect(db.engine)
            
            # Check kas_mutasi table
            if 'kas_mutasi' in inspector.get_table_names():
                kas_columns = {col['name']: col['type'] for col in inspector.get_columns('kas_mutasi')}
                
                if 'sumber_saldo' not in kas_columns:
                    print('[DB-MIGRATE] Adding sumber_saldo to kas_mutasi...')
                    with db.engine.connect() as conn:
                        conn.execute(text('ALTER TABLE kas_mutasi ADD COLUMN sumber_saldo VARCHAR(20) DEFAULT "harian"'))
                        conn.commit()
                    print('[DB-MIGRATE] ✓ sumber_saldo added')
            
            # Check daily_saldo_archive table
            if 'daily_saldo_archive' in inspector.get_table_names():
                archive_columns = {col['name']: col['type'] for col in inspector.get_columns('daily_saldo_archive')}
                
                if 'saldo_harian_input' not in archive_columns:
                    print('[DB-MIGRATE] Adding saldo_harian_input to daily_saldo_archive...')
                    with db.engine.connect() as conn:
                        conn.execute(text('ALTER TABLE daily_saldo_archive ADD COLUMN saldo_harian_input FLOAT DEFAULT 0'))
                        conn.commit()
                    print('[DB-MIGRATE] ✓ saldo_harian_input added')
            
            print('[DB-MIGRATE] Database columns verified')
    except Exception as e:
        print(f'[DB-MIGRATE] Warning: {e}')
        pass

def start_scheduler():
    """Start background scheduler."""
    if not SCHEDULER_AVAILABLE or scheduler is None:
        print('[SCHEDULER] Skipped (APScheduler not installed)')
        return

    if not scheduler.running:
        # Schedule daily report generation at 21:30 (9:30 PM)
        scheduler.add_job(
            generate_daily_report,
            trigger=CronTrigger(hour=21, minute=30),
            id='daily_report_generate',
            name='Generate daily financial report',
            replace_existing=True
        )
        # Schedule daily archive at 22:30 (10:30 PM)
        scheduler.add_job(
            archive_daily_saldo,
            trigger=CronTrigger(hour=22, minute=30),
            id='daily_saldo_archive',
            name='Archive daily saldo',
            replace_existing=True
        )
        scheduler.start()
        print('[SCHEDULER] Daily report scheduler started (21:30 every day)')
        print('[SCHEDULER] Daily saldo archive scheduler started (22:30 every day)')

@app.before_request
def _auto_start_telegram_bot() -> None:
    _start_telegram_bot_if_configured(app)

# ==================== MEMBER CONFIG ====================

POINTS_PER_RUPIAH = 10000  # 1 point per Rp 10.000
LEVEL_RULES = [
    (0, 'Bronze'),
    (1000, 'Silver'),
    (5000, 'Gold'),
]

def get_member_level(points):
    level = 'Bronze'
    for min_points, name in LEVEL_RULES:
        if points >= min_points:
            level = name
    return level

def calculate_points_from_total(total_rupiah):
    if total_rupiah <= 0:
        return 0
    return int(total_rupiah // POINTS_PER_RUPIAH)

# ==================== TEMPLATE FILTERS ====================

@app.template_filter('format_datetime')
def format_datetime_filter(dt_obj):
    """Format datetime untuk display (no timezone conversion, langsung tampilkan)"""
    if dt_obj is None:
        return None
    # Langsung return datetime object as-is untuk di-format di template
    return dt_obj

@app.context_processor
def inject_globals():
    return {
        'current_time': get_local_now(),
        'timezone_name': get_local_timezone_name()
    }

@app.route('/img/<path:filename>')
def serve_img(filename):
    return send_from_directory(os.path.join(base_app_dir, 'img'), filename)

# ==================== REQUEST HANDLERS ====================

@app.before_request
def check_navigation():
    # Force HTTP for all requests (prevent HTTPS auto-upgrade)
    if request.headers.get('X-Forwarded-Proto') == 'https':
        return redirect(request.url.replace('https://', 'http://'), code=301)
    
    if current_user.is_authenticated:
        # If user tries to access login page while logged in, redirect to index
        # But allow register page for admin to add new users
        if request.endpoint == 'login':
            return redirect(url_for('index'))
    else:
        # If not logged in, only allow login/register and public assets
        if request.endpoint and request.endpoint not in ['login', 'register', 'static', 'serve_img']:
            return redirect(url_for('login'))

@app.after_request
def add_security_headers(response):
    # Relaxed CSP for development - allow more sources
    response.headers['Content-Security-Policy'] = "default-src 'self' 'unsafe-inline' 'unsafe-eval' data: blob:; script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://code.jquery.com https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com https://cdnjs.cloudflare.com; font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; img-src 'self' data: https://*; connect-src 'self' https://*; object-src 'none'; base-uri 'self'; form-action 'self'"
    # Disable HSTS for development
    response.headers['Strict-Transport-Security'] = 'max-age=0'
    # Prevent caching for authenticated or protected pages to avoid back-button access after logout
    if request.endpoint and request.endpoint not in ['static', 'login']:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nama = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='kasir')
    
    def set_password(self, password):
        """Set password dengan hashing yang lebih aman"""
        self.password_hash = generate_password_hash(
            password, 
            method='pbkdf2:sha256', 
            salt_length=16
        )
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    @staticmethod
    def validate_password_strength(password):
        """Validasi kekuatan password"""
        if len(password) < 8:
            return False, "Password minimal 8 karakter"
        if not any(c.isupper() for c in password):
            return False, "Password harus mengandung huruf besar"
        if not any(c.isdigit() for c in password):
            return False, "Password harus mengandung angka"
        return True, "Password valid"

class Kategori(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False, unique=True)
    deskripsi = db.Column(db.Text)
    produk = db.relationship('Produk', backref='kategori_ref', lazy=True)

class Member(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False)
    no_telp = db.Column(db.String(30))
    alamat = db.Column(db.Text)
    catatan = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=get_local_now)
    points = db.Column(db.Integer, default=0)
    total_spent = db.Column(db.Float, default=0)

    def get_level(self):
        return get_member_level(self.points or 0)

class Produk(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kode = db.Column(db.String(50), unique=True, nullable=False)
    nama = db.Column(db.String(200), nullable=False)
    deskripsi = db.Column(db.Text)
    harga_beli = db.Column(db.Float, nullable=False)
    harga_jual = db.Column(db.Float, nullable=False)
    stok = db.Column(db.Integer, default=0)
    kategori_id = db.Column(db.Integer, db.ForeignKey('kategori.id'))
    minimal_stok = db.Column(db.Integer, default=5)
    satuan = db.Column(db.String(20), default='pcs')
    harga_variasi = db.relationship('HargaVariasi', backref='produk', lazy=True, cascade='all, delete-orphan', order_by='HargaVariasi.min_qty')
    varian_produk = db.relationship('VarianProduk', backref='produk_ref', lazy=True, cascade='all, delete-orphan', order_by='VarianProduk.created_at')
    
    def get_harga_by_qty(self, qty):
        """Dapatkan harga berdasarkan quantity"""
        # Cek apakah ada harga variasi
        if self.harga_variasi:
            # Sort descending by min_qty untuk cek dari qty terbesar
            for variant in reversed(self.harga_variasi):
                if qty >= variant.min_qty:
                    return variant.harga
        # Default ke harga jual
        return self.harga_jual
    
    def to_dict(self):
        # Get price variants
        variants = []
        if self.harga_variasi:
            variants = [{'min_qty': v.min_qty, 'harga': v.harga} for v in self.harga_variasi]
            
        # Get product variants
        product_variants = []
        if self.varian_produk:
            product_variants = [{
                'id': v.id,
                'nama_varian': v.nama_varian,
                'barcode_varian': v.barcode_varian,
                'stok': v.stok
            } for v in self.varian_produk]
            
        return {
            'id': self.id,
            'kode': self.kode,
            'nama': self.nama,
            'harga_jual': self.harga_jual,
            'harga_variasi': variants,
            'varian_produk': product_variants,
            'stok': self.stok,
            'satuan': self.satuan
        }

class HargaVariasi(db.Model):
    """Model untuk menyimpan harga bertingkat berdasarkan quantity"""
    id = db.Column(db.Integer, primary_key=True)
    produk_id = db.Column(db.Integer, db.ForeignKey('produk.id'), nullable=False)
    min_qty = db.Column(db.Integer, nullable=False)  # Minimal quantity untuk harga ini
    harga = db.Column(db.Float, nullable=False)  # Harga per unit
    keterangan = db.Column(db.String(100))  # Opsional: keterangan tier harga

class VarianProduk(db.Model):
    """Model untuk menyimpan varian produk dengan barcode"""
    id = db.Column(db.Integer, primary_key=True)
    produk_id = db.Column(db.Integer, db.ForeignKey('produk.id'), nullable=False)
    nama_varian = db.Column(db.String(200), nullable=False)  # Nama varian (misal: "Varian A", "Kemasan Besar")
    barcode_varian = db.Column(db.String(100), unique=True, nullable=False)  # Barcode unik untuk varian
    stok = db.Column(db.Integer, default=0)  # Stok untuk varian ini
    created_at = db.Column(db.DateTime, default=get_local_now)
    updated_at = db.Column(db.DateTime, default=get_local_now, onupdate=get_local_now)

class Transaksi(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kode_transaksi = db.Column(db.String(50), unique=True, nullable=False)
    tanggal = db.Column(db.DateTime, default=get_local_now)
    subtotal = db.Column(db.Float, nullable=False, default=0)
    discount_percent = db.Column(db.Float, default=0)
    discount_amount = db.Column(db.Float, default=0)
    total = db.Column(db.Float, nullable=False)
    bayar = db.Column(db.Float, nullable=False)
    kembalian = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), default='tunai')  # tunai, qris, ewallet, debit, hutang
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'))
    member_manual = db.Column(db.String(100))  # Untuk input manual nama/telp member
    points_earned = db.Column(db.Integer, default=0)
    user = db.relationship('User', backref='transaksi')
    member = db.relationship('Member', backref='transaksi')
    items = db.relationship('TransaksiItem', backref='transaksi_ref', lazy=True)

class TransaksiItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    transaksi_id = db.Column(db.Integer, db.ForeignKey('transaksi.id'))
    produk_id = db.Column(db.Integer, db.ForeignKey('produk.id'))
    produk = db.relationship('Produk')
    jumlah = db.Column(db.Integer, nullable=False)
    harga = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    varian_barcode = db.Column(db.String(100))  # Barcode varian jika ada
    varian_nama = db.Column(db.String(200))     # Nama varian jika ada



class Pengaturan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Text)
    
    @staticmethod
    def get(key, default=''):
        setting = Pengaturan.query.filter_by(key=key).first()
        return setting.value if setting else default
    
    @staticmethod
    def set(key, value):
        setting = Pengaturan.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = Pengaturan(key=key, value=value)
            db.session.add(setting)
        db.session.commit()



# ==================== FORMS ====================

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

class RegisterForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=80)])
    nama = StringField('Nama Lengkap', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Konfirmasi Password', validators=[DataRequired()])
    role = SelectField('Role', choices=[('kasir', 'Kasir'), ('admin', 'Admin')])
    submit = SubmitField('Register')
    
    def validate_username(self, username):
        """Validasi username"""
        username_val = username.data
        if ' ' in username_val:
            raise ValidationError('Username tidak boleh mengandung spasi')
        if len(username_val) < 3:
            raise ValidationError('Username minimal 3 karakter')
        
        user = User.query.filter_by(username=username_val).first()
        if user:
            raise ValidationError('Username sudah digunakan.')
    
    def validate_password(self, field):
        """Validasi kekuatan password"""
        password = field.data
        if len(password) < 8:
            raise ValidationError('Password minimal 8 karakter')
        if not any(c.isupper() for c in password):
            raise ValidationError('Password harus mengandung huruf besar')
        if not any(c.isdigit() for c in password):
            raise ValidationError('Password harus mengandung angka')
    
    def validate_confirm_password(self, confirm_password):
        if self.password.data != confirm_password.data:
            raise ValidationError('Password tidak cocok.')

class ProdukForm(FlaskForm):
    product_id = HiddenField()
    kode = StringField('Kode Produk', validators=[DataRequired()])
    nama = StringField('Nama Produk', validators=[DataRequired()])
    deskripsi = TextAreaField('Deskripsi')
    harga_beli = DecimalField('Harga Beli', validators=[DataRequired(), NumberRange(min=0)])
    harga_jual = DecimalField('Harga Jual', validators=[DataRequired(), NumberRange(min=0)])
    stok = IntegerField('Stok Awal', validators=[DataRequired(), NumberRange(min=0)])
    kategori_id = SelectField('Kategori', coerce=int, validators=[Optional()])
    minimal_stok = IntegerField('Minimal Stok', validators=[DataRequired(), NumberRange(min=0)])
    satuan = StringField('Satuan', validators=[DataRequired()])
    submit = SubmitField('Simpan')
    
    def validate_kode(self, kode):
        produk = Produk.query.filter_by(kode=kode.data).first()
        if produk:
            # Allow same kode when editing the current product
            if self.product_id.data:
                try:
                    if produk.id == int(self.product_id.data):
                        return
                except ValueError:
                    pass
            raise ValidationError('Kode produk sudah digunakan! Gunakan kode yang berbeda.')

class KategoriForm(FlaskForm):
    nama = StringField('Nama Kategori', validators=[DataRequired()])
    deskripsi = TextAreaField('Deskripsi')
    submit = SubmitField('Simpan')

class MemberForm(FlaskForm):
    nama = StringField('Nama Member', validators=[DataRequired(), Length(max=100)])
    no_telp = StringField('Nomor Telepon', validators=[Optional(), Length(max=30)])
    alamat = TextAreaField('Alamat', validators=[Optional(), Length(max=500)])
    catatan = TextAreaField('Catatan', validators=[Optional(), Length(max=500)])
    submit = SubmitField('Simpan')

# ==================== USER LOADER ====================

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# ==================== ROUTES ====================

@app.route('/')
@login_required
def index():
    total_produk = Produk.query.count()
    today = date.today()
    total_transaksi_hari_ini = Transaksi.query.filter(db.func.date(Transaksi.tanggal) == today).count()
    produk_habis = Produk.query.filter(Produk.stok <= Produk.minimal_stok).count()
    
    return render_template('index.html', 
                         total_produk=total_produk,
                         total_transaksi=total_transaksi_hari_ini,
                         produk_habis=produk_habis,
                         current_time=get_local_now(),
                         timezone_name=get_local_timezone_name())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    # Ensure CSRF token is generated and stored in session for the login form.
    generate_csrf()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Login berhasil!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Username atau password salah!', 'danger')
    
    return render_template('auth/login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            nama=form.nama.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'User {form.username.data} berhasil ditambahkan!', 'success')
        return redirect(url_for('register'))
    
    # Get all users to display in table
    users = User.query.order_by(User.id).all()
    
    return render_template('auth/register.html', form=form, users=users)

@app.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'}), 403
    
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'Tidak bisa menghapus akun sendiri!'}), 400
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'User {username} berhasil dihapus!'})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    response = redirect(url_for('login'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ==================== PRODUK ROUTES ====================

@app.route('/produk')
@login_required
def list_produk():
    filter_stok = request.args.get('filter', 'semua')
    search = request.args.get('q', '').strip()
    kategori_id = request.args.get('kategori_id', '').strip()
    page, per_page = _get_pagination_args(default_per_page=50, max_per_page=200)
    
    query = Produk.query
    
    if filter_stok == 'habis':
        query = query.filter(Produk.stok == 0)
    elif filter_stok == 'hampir_habis':
        query = query.filter(Produk.stok > 0, Produk.stok <= Produk.minimal_stok)
    elif filter_stok == 'tersedia':
        query = query.filter(Produk.stok > 0)

    kategori_filter = None
    if kategori_id.isdigit():
        kategori_filter = Kategori.query.get(int(kategori_id))
        if kategori_filter:
            query = query.filter(Produk.kategori_id == kategori_filter.id)

    if search:
        # Cari di produk utama (nama, kode) atau di barcode varian
        query = query.filter(
            or_(
                Produk.nama.ilike(f'%{search}%'),
                Produk.kode.ilike(f'%{search}%'),
                # Cari produk yang memiliki varian dengan barcode yang cocok
                Produk.id.in_(
                    db.session.query(VarianProduk.produk_id)
                    .filter(VarianProduk.barcode_varian.ilike(f'%{search}%'))
                )
            )
        )
    
    pagination = _paginate_query(query, page=page, per_page=per_page, order_by=Produk.kode)
    produk_list = pagination['items']
    kategori_list = Kategori.query.all()
    return render_template('produk/list.html', 
                         produk_list=produk_list,
                         kategori_list=kategori_list,
                         filter_stok=filter_stok,
                         search=search,
                         kategori_filter=kategori_filter,
                         pagination=pagination)

@app.route('/produk/tambah', methods=['GET', 'POST'])
@login_required
def tambah_produk():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    form = ProdukForm()
    form.kategori_id.choices = [(k.id, k.nama) for k in Kategori.query.all()]
    
    if form.validate_on_submit():
        try:
            produk = Produk(
                kode=form.kode.data,
                nama=form.nama.data,
                deskripsi=form.deskripsi.data,
                harga_beli=float(form.harga_beli.data),
                harga_jual=float(form.harga_jual.data),
                stok=form.stok.data,
                kategori_id=form.kategori_id.data if form.kategori_id.data else None,
                minimal_stok=form.minimal_stok.data,
                satuan=form.satuan.data
            )
            db.session.add(produk)
            db.session.flush()  # Get produk.id
            
            # Handle harga variasi (tier pricing)
            variant_min_qtys = request.form.getlist('variant_min_qty[]')
            variant_hargas = request.form.getlist('variant_harga[]')
            variant_keterangans = request.form.getlist('variant_keterangan[]')
            
            for i in range(len(variant_min_qtys)):
                if variant_min_qtys[i] and variant_hargas[i]:
                    harga_variasi = HargaVariasi(
                        produk_id=produk.id,
                        min_qty=int(variant_min_qtys[i]),
                        harga=float(variant_hargas[i]),
                        keterangan=variant_keterangans[i] if i < len(variant_keterangans) else None
                    )
                    db.session.add(harga_variasi)
            
            # Handle varian barang (product variants with barcode)
            varian_namas = request.form.getlist('varian_nama[]')
            varian_barcodes = request.form.getlist('varian_barcode[]')
            varian_stoks = request.form.getlist('varian_stok[]')
            
            for i in range(len(varian_namas)):
                if varian_namas[i] and varian_barcodes[i]:
                    # Check if barcode sudah ada
                    existing_varian = VarianProduk.query.filter_by(barcode_varian=varian_barcodes[i]).first()
                    if existing_varian and existing_varian.produk_id != produk.id:
                        flash(f'Barcode {varian_barcodes[i]} sudah digunakan di produk lain!', 'danger')
                        db.session.rollback()
                        return render_template('produk/form.html', form=form, title='Tambah Produk', produk=None)
                    
                    # Get stok value, default to 0 if not provided or invalid
                    try:
                        stok_varian = int(varian_stoks[i]) if i < len(varian_stoks) and varian_stoks[i] else 0
                    except (ValueError, TypeError):
                        stok_varian = 0
                    
                    varian_produk = VarianProduk(
                        produk_id=produk.id,
                        nama_varian=varian_namas[i],
                        barcode_varian=varian_barcodes[i],
                        stok=stok_varian
                    )
                    db.session.add(varian_produk)
            
            db.session.commit()
            flash('Produk berhasil ditambahkan!', 'success')
            return redirect(url_for('list_produk'))
        except ValueError as e:
            db.session.rollback()
            flash(f'Error: Input tidak valid - {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan produk: {str(e)}', 'danger')
    
    return render_template('produk/form.html', form=form, title='Tambah Produk', produk=None)

@app.route('/produk/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_produk(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    produk = Produk.query.get_or_404(id)
    form = ProdukForm(obj=produk)
    form.product_id.data = produk.id
    form.kategori_id.choices = [(k.id, k.nama) for k in Kategori.query.all()]
    
    if form.validate_on_submit():
        try:
            produk.kode = form.kode.data
            produk.nama = form.nama.data
            produk.deskripsi = form.deskripsi.data
            produk.harga_beli = float(form.harga_beli.data)
            produk.harga_jual = float(form.harga_jual.data)
            produk.stok = form.stok.data
            produk.kategori_id = form.kategori_id.data if form.kategori_id.data else None
            produk.minimal_stok = form.minimal_stok.data
            produk.satuan = form.satuan.data
            
            # Update harga variasi - hapus yang lama, insert yang baru
            # Karena relationship cascade delete-orphan, akan auto-delete
            produk.harga_variasi.clear()
            
            variant_min_qtys = request.form.getlist('variant_min_qty[]')
            variant_hargas = request.form.getlist('variant_harga[]')
            variant_keterangans = request.form.getlist('variant_keterangan[]')
            
            for i in range(len(variant_min_qtys)):
                if variant_min_qtys[i] and variant_hargas[i]:
                    harga_variasi = HargaVariasi(
                        produk_id=produk.id,
                        min_qty=int(variant_min_qtys[i]),
                        harga=float(variant_hargas[i]),
                        keterangan=variant_keterangans[i] if i < len(variant_keterangans) else None
                    )
                    db.session.add(harga_variasi)
            
            # Update varian barang - hapus yang lama, insert yang baru
            produk.varian_produk.clear()
            
            varian_namas = request.form.getlist('varian_nama[]')
            varian_barcodes = request.form.getlist('varian_barcode[]')
            varian_stoks = request.form.getlist('varian_stok[]')
            
            for i in range(len(varian_namas)):
                if varian_namas[i] and varian_barcodes[i]:
                    # Check if barcode sudah ada di produk lain
                    existing_varian = VarianProduk.query.filter_by(barcode_varian=varian_barcodes[i]).first()
                    if existing_varian and existing_varian.produk_id != produk.id:
                        flash(f'Barcode {varian_barcodes[i]} sudah digunakan di produk lain!', 'danger')
                        db.session.rollback()
                        return render_template('produk/form.html', form=form, title='Edit Produk', produk=produk)
                    
                    # Get stok value, default to 0 if not provided or invalid
                    try:
                        stok_varian = int(varian_stoks[i]) if i < len(varian_stoks) and varian_stoks[i] else 0
                    except (ValueError, TypeError):
                        stok_varian = 0
                    
                    varian_produk = VarianProduk(
                        produk_id=produk.id,
                        nama_varian=varian_namas[i],
                        barcode_varian=varian_barcodes[i],
                        stok=stok_varian
                    )
                    db.session.add(varian_produk)
            
            db.session.commit()
            flash('Produk berhasil diperbarui!', 'success')
            return redirect(url_for('list_produk'))
        except ValueError as e:
            db.session.rollback()
            flash(f'Error: Input tidak valid - {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan produk: {str(e)}', 'danger')
    
    return render_template('produk/form.html', form=form, title='Edit Produk', produk=produk)

@app.route('/produk/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_produk(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    produk = Produk.query.get_or_404(id)
    db.session.delete(produk)
    db.session.commit()
    flash('Produk berhasil dihapus!', 'success')
    return redirect(url_for('list_produk'))

# ==================== KATEGORI ROUTES ====================

@app.route('/kategori')
@login_required
def list_kategori():
    kategori_list = Kategori.query.all()
    return render_template('produk/kategori.html', kategori_list=kategori_list)

@app.route('/kategori/tambah', methods=['GET', 'POST'])
@login_required
def tambah_kategori():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    form = KategoriForm()
    if form.validate_on_submit():
        try:
            kategori = Kategori(nama=form.nama.data, deskripsi=form.deskripsi.data)
            db.session.add(kategori)
            db.session.commit()
            flash('Kategori berhasil ditambahkan!', 'success')
            return redirect(url_for('list_kategori'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan kategori: {str(e)}', 'danger')
    
    return render_template('produk/kategori_form.html', form=form, title='Tambah Kategori')

@app.route('/kategori/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_kategori(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    kategori = Kategori.query.get_or_404(id)
    form = KategoriForm(obj=kategori)
    
    if form.validate_on_submit():
        try:
            kategori.nama = form.nama.data
            kategori.deskripsi = form.deskripsi.data
            db.session.commit()
            flash('Kategori berhasil diperbarui!', 'success')
            return redirect(url_for('list_kategori'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui kategori: {str(e)}', 'danger')
    
    return render_template('produk/kategori_form.html', form=form, title='Edit Kategori')

@app.route('/kategori/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_kategori(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    kategori = Kategori.query.get_or_404(id)
    db.session.delete(kategori)
    db.session.commit()
    flash('Kategori berhasil dihapus!', 'success')
    return redirect(url_for('list_kategori'))

@app.route('/kategori/<int:id>/produk')
@login_required
def kategori_produk(id):
    kategori = Kategori.query.get_or_404(id)
    page, per_page = _get_pagination_args(default_per_page=50, max_per_page=200)
    query = (Produk.query
        .filter(Produk.kategori_id == kategori.id)
        .options(joinedload(Produk.harga_variasi)))
    pagination = _paginate_query(query, page=page, per_page=per_page, order_by=Produk.kode)
    produk_list = pagination['items']
    return render_template('produk/kategori_produk.html',
                         kategori=kategori,
                         produk_list=produk_list,
                         pagination=pagination)

# ==================== MEMBER ROUTES ====================

@app.route('/member')
@login_required
def list_member():
    search = request.args.get('q', '').strip()
    page, per_page = _get_pagination_args(default_per_page=50, max_per_page=200)
    query = Member.query

    if search:
        query = query.filter(
            or_(
                Member.nama.ilike(f'%{search}%'),
                Member.no_telp.ilike(f'%{search}%')
            )
        )

    pagination = _paginate_query(query, page=page, per_page=per_page, order_by=Member.nama)
    member_list = pagination['items']
    return render_template('member/list.html', member_list=member_list, search=search, pagination=pagination)

@app.route('/member/<int:id>/transaksi')
@login_required
def member_transaksi(id):
    member = Member.query.get_or_404(id)
    page, per_page = _get_pagination_args(default_per_page=50, max_per_page=200)
    query = (
        Transaksi.query
        .filter_by(member_id=member.id)
    )
    total_transaksi = query.order_by(None).count()
    total_belanja = (db.session.query(db.func.coalesce(db.func.sum(Transaksi.total), 0))
                     .filter(Transaksi.member_id == member.id)
                     .scalar()) or 0
    pagination = _paginate_query(query, page=page, per_page=per_page, order_by=Transaksi.tanggal.desc())
    transaksi_list = pagination['items']
    return render_template(
        'member/transaksi.html',
        member=member,
        transaksi_list=transaksi_list,
        total_transaksi=total_transaksi,
        total_belanja=total_belanja,
        pagination=pagination
    )

@app.route('/member/export')
@login_required
def export_member():
    members = Member.query.order_by(Member.nama).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for member in members:
        ws.append([
            member.nama,
            member.no_telp or '',
            member.alamat or '',
            member.catatan or '',
            member.points or 0,
            member.total_spent or 0
        ])

    # Ranking sheet
    ws_rank = wb.create_sheet('Ranking')
    ws_rank.append(['Member', 'Total Spent'])
    for cell in ws_rank[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    top_members = sorted(members, key=lambda m: m.total_spent or 0, reverse=True)[:10]
    for member in top_members:
        ws_rank.append([member.nama, member.total_spent or 0])

    if top_members:
        data = Reference(ws_rank, min_col=2, min_row=1, max_row=len(top_members) + 1)
        categories = Reference(ws_rank, min_col=1, min_row=2, max_row=len(top_members) + 1)
        chart = BarChart()
        chart.title = 'Top Member by Total Spent'
        chart.y_axis.title = 'Total Spent'
        chart.x_axis.title = 'Member'
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        ws_rank.add_chart(chart, 'D2')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=member_export.xlsx'
    return response

@app.route('/member/template')
@login_required
def download_member_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws.append(['Siti Aminah', '081234567890', 'Jl. Melati No. 1', 'Pelanggan loyal', 0, 0])
    ws.append(['Budi Santoso', '082233445566', 'Jl. Kenanga No. 12', 'Suka belanja grosir', 120, 450000])
    ws.append(['Rina', '', 'Pasar Blok A', 'Tanpa no telp', 30, 125000])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=member_template.xlsx'
    return response

@app.route('/member/import', methods=['POST'])
@login_required
def import_member():
    if 'file' not in request.files:
        flash('File XLSX tidak ditemukan.', 'danger')
        return redirect(url_for('list_member'))

    file = request.files['file']
    if not file or file.filename == '':
        flash('File XLSX tidak ditemukan.', 'danger')
        return redirect(url_for('list_member'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb['Members'] if 'Members' in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('File XLSX kosong.', 'danger')
            return redirect(url_for('list_member'))

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        required_columns = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
        if not set(required_columns).issubset(set(headers)):
            flash('Header XLSX tidak sesuai. Gunakan template yang disediakan.', 'danger')
            return redirect(url_for('list_member'))

        col_index = {name: headers.index(name) for name in required_columns}
        added = 0
        updated = 0

        for row in rows[1:]:
            nama = (row[col_index['nama']] or '').strip() if row[col_index['nama']] else ''
            if not nama:
                continue

            no_telp_val = row[col_index['no_telp']]
            alamat_val = row[col_index['alamat']]
            catatan_val = row[col_index['catatan']]

            no_telp = str(no_telp_val).strip() if no_telp_val is not None else None
            alamat = str(alamat_val).strip() if alamat_val is not None else None
            catatan = str(catatan_val).strip() if catatan_val is not None else None
            no_telp = no_telp or None
            alamat = alamat or None
            catatan = catatan or None

            try:
                points = int(float(row[col_index['points']] or 0))
            except (TypeError, ValueError):
                points = 0

            try:
                total_spent = float(row[col_index['total_spent']] or 0)
            except (TypeError, ValueError):
                total_spent = 0

            if no_telp:
                existing = Member.query.filter_by(no_telp=no_telp).first()
            else:
                existing = Member.query.filter_by(nama=nama).first()
            if existing:
                existing.alamat = alamat
                existing.catatan = catatan
                existing.points = points
                existing.total_spent = total_spent
                updated += 1
            else:
                member = Member(
                    nama=nama,
                    no_telp=no_telp,
                    alamat=alamat,
                    catatan=catatan,
                    points=points,
                    total_spent=total_spent
                )
                db.session.add(member)
                added += 1

        db.session.commit()
        flash(f'Import selesai: {added} ditambah, {updated} diperbarui.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal import member: {str(e)}', 'danger')

    return redirect(url_for('list_member'))

@app.route('/member/tambah', methods=['GET', 'POST'])
@login_required
def tambah_member():
    form = MemberForm()

    if form.validate_on_submit():
        try:
            member = Member(
                nama=form.nama.data.strip(),
                no_telp=form.no_telp.data.strip() if form.no_telp.data else None,
                alamat=form.alamat.data.strip() if form.alamat.data else None,
                catatan=form.catatan.data.strip() if form.catatan.data else None
            )
            db.session.add(member)
            db.session.commit()
            flash('Member berhasil ditambahkan!', 'success')
            return redirect(url_for('list_member'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan member: {str(e)}', 'danger')

    return render_template('member/form.html', form=form, title='Tambah Member')

@app.route('/member/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_member(id):
    member = Member.query.get_or_404(id)
    form = MemberForm(obj=member)

    if form.validate_on_submit():
        try:
            member.nama = form.nama.data.strip()
            member.no_telp = form.no_telp.data.strip() if form.no_telp.data else None
            member.alamat = form.alamat.data.strip() if form.alamat.data else None
            member.catatan = form.catatan.data.strip() if form.catatan.data else None
            db.session.commit()
            flash('Member berhasil diperbarui!', 'success')
            return redirect(url_for('list_member'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui member: {str(e)}', 'danger')

    return render_template('member/form.html', form=form, title='Edit Member')

@app.route('/member/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_member(id):
    member = Member.query.get_or_404(id)
    try:
        Transaksi.query.filter_by(member_id=member.id).update(
            {Transaksi.member_id: None},
            synchronize_session=False
        )
        db.session.delete(member)
        db.session.commit()
        flash('Member berhasil dihapus!', 'success')
        return redirect(url_for('list_member'))
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus member: {str(e)}', 'danger')
        return redirect(url_for('list_member'))

# ==================== TRANSAKSI ROUTES ====================

@app.route('/kasir')
@login_required
def kasir():
    kategori_list = Kategori.query.all()
    return render_template('transaksi/kasir.html', 
                         kategori_list=kategori_list,
                         produk_list=[],
                         member_list=[])

@app.route('/api/produk')
@login_required
def get_api_produk():
    search = request.args.get('search', '').strip()
    kategori_id = request.args.get('kategori_id', type=int)
    limit = request.args.get('limit', 50, type=int) or 50
    limit = max(1, min(200, limit))
    
    query = Produk.query.filter(
        or_(
            Produk.stok > 0,
            Produk.id.in_(
                db.session.query(VarianProduk.produk_id)
                .filter(VarianProduk.stok > 0)
            )
        )
    )
    
    if search:
        query = query.filter(
            or_(
                Produk.nama.ilike(f'%{search}%'),
                Produk.kode.ilike(f'%{search}%'),
                # Cari produk yang memiliki varian dengan barcode yang cocok
                Produk.id.in_(
                    db.session.query(VarianProduk.produk_id)
                    .filter(VarianProduk.barcode_varian.ilike(f'%{search}%'))
                )
            )
        )
    
    if kategori_id:
        query = query.filter_by(kategori_id=kategori_id)
    
    produk_list = query.order_by(Produk.nama).limit(limit).all()  # Batasi hasil untuk performa
    
    return jsonify([p.to_dict() for p in produk_list])


@app.route('/api/member')
@login_required
def get_api_member():
    search = (request.args.get('search', '') or '').strip()
    limit = request.args.get('limit', 10, type=int) or 10
    limit = max(1, min(50, limit))

    query = Member.query
    if search:
        if search.isdigit() and len(search) == 4:
            query = query.filter(Member.no_telp.ilike(f'%{search}'))
        else:
            query = query.filter(
                or_(
                    Member.nama.ilike(f'%{search}%'),
                    Member.no_telp.ilike(f'%{search}%')
                )
            )

    members = query.order_by(Member.nama).limit(limit).all()
    return jsonify([
        {
            'id': m.id,
            'nama': m.nama,
            'no_telp': m.no_telp or ''
        }
        for m in members
    ])

@app.route('/api/varian/<int:varian_id>/stok', methods=['POST'])
@login_required
@csrf.exempt
def update_varian_stok(varian_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak'}), 403
    
    try:
        data = request.get_json()
        if not data or 'stok' not in data:
            return jsonify({'success': False, 'message': 'Data stok tidak valid'}), 400
        
        stok_baru = int(data['stok'])
        if stok_baru < 0:
            return jsonify({'success': False, 'message': 'Stok tidak boleh negatif'}), 400
        
        varian = VarianProduk.query.get_or_404(varian_id)
        varian.stok = stok_baru
        varian.updated_at = get_local_now()
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Stok varian {varian.nama_varian} berhasil diupdate ke {stok_baru}'
        })
        
    except ValueError:
        return jsonify({'success': False, 'message': 'Stok harus berupa angka'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Gagal update stok: {str(e)}'}), 500

@app.route('/transaksi/checkout', methods=['POST'])
@login_required
@csrf.exempt
def checkout():
    print("\n" + "="*50)
    print("CHECKOUT PROCESS STARTED")
    print("="*50)
    
    try:
        # Log user dan IP
        user_info = f"User: {current_user.username} (ID: {current_user.id})"
        ip_info = f"IP: {request.remote_addr}"
        print(f"[Checkout] {user_info} | {ip_info}")
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Data tidak valid'})
        
        items = data.get('items', [])
        total = data.get('total', 0)
        bayar = data.get('bayar', 0)
        payment_method = data.get('payment_method', 'tunai')
        member_id = data.get('member_id')
        member_manual = data.get('member_manual')  # Input manual nama/telp
        
        print(f"[Checkout] Items: {len(items)}")
        print(f"[Checkout] Total: Rp {total}")
        print(f"[Checkout] Bayar: Rp {bayar}")
        
        if not items:
            return jsonify({'success': False, 'message': 'Keranjang kosong'})
        
        # Validasi tipe data
        try:
            total = float(total)
            bayar = float(bayar)
        except ValueError:
            return jsonify({'success': False, 'message': 'Format angka tidak valid'})
        
        member = None
        if member_id:
            try:
                member = Member.query.get(int(member_id))
            except (TypeError, ValueError):
                return jsonify({'success': False, 'message': 'Member tidak valid'})
            if not member:
                return jsonify({'success': False, 'message': 'Member tidak ditemukan'})

        # Calculate totals server-side
        subtotal = 0
        for item in items:
            try:
                qty = int(item.get('quantity', 1))
                price = float(item.get('price', 0))
            except (TypeError, ValueError):
                return jsonify({'success': False, 'message': 'Format item tidak valid'})
            if qty < 1 or price < 0:
                return jsonify({'success': False, 'message': 'Data item tidak valid'})
            subtotal += qty * price

        total = subtotal
        points_earned = calculate_points_from_total(total) if member else 0

        if bayar < total:
            return jsonify({'success': False, 'message': 'Pembayaran kurang'})
        
        kode_transaksi = f'TRX{get_local_now().strftime("%Y%m%d%H%M%S")}'
        print(f"[Checkout] Transaction code: {kode_transaksi}")
        print(f"[Checkout] Local Time: {get_local_now().strftime('%Y-%m-%d %H:%M:%S')} {get_local_timezone_name()}")
        
        # Buat transaksi
        transaksi = Transaksi(
            kode_transaksi=kode_transaksi,
            subtotal=subtotal,
            discount_percent=0,
            discount_amount=0,
            total=total,
            bayar=bayar,
            kembalian=bayar - total,
            payment_method=payment_method,
            user_id=current_user.id,
            member_id=member.id if member else None,
            member_manual=member_manual if not member else None,  # Simpan input manual jika bukan member terdaftar
            points_earned=points_earned
        )
        db.session.add(transaksi)
        db.session.flush()
        
        # Log timestamp transaksi
        print(f"[Checkout] Transaction timestamp: {transaksi.tanggal.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Tambahkan items
        for item in items:
            produk = Produk.query.get(item['id'])
            if not produk:
                return jsonify({'success': False, 'message': f'Produk tidak ditemukan'})
                
            quantity = item.get('quantity', 1)
            scanned_variant = item.get('scanned_variant')
            
            # Cek stok berdasarkan varian atau produk utama
            if scanned_variant and scanned_variant.get('barcode'):
                # Cari varian berdasarkan barcode
                varian = VarianProduk.query.filter_by(
                    produk_id=item['id'], 
                    barcode_varian=scanned_variant['barcode']
                ).first()
                
                if not varian:
                    return jsonify({'success': False, 'message': f'Varian produk tidak ditemukan'})
                
                if varian.stok < quantity:
                    return jsonify({'success': False, 'message': f'Stok varian {scanned_variant.get("nama", "Unknown")} tidak cukup'})
                
                # Kurangi stok varian
                varian.stok -= quantity
                item_name = f"{produk.nama} - {scanned_variant.get('nama', 'Varian')}"
                print(f"[Checkout] Item: {item_name} x{quantity} (varian)")
                
                # Check low stock for variant
                if TELEGRAM_AVAILABLE:
                    try:
                        bot = get_telegram_bot()
                        threshold = app.config.get('TELEGRAM_NOTIFY_LOW_STOCK_THRESHOLD', 10)
                        if bot and varian.stok <= threshold:
                            bot.notify_low_stock_sync(
                                produk_nama=f"{produk.nama} - {scanned_variant.get('nama', 'Varian')}",
                                stok=varian.stok,
                                kategori=produk.kategori.nama if produk.kategori else "Tanpa Kategori"
                            )
                            print(f"[Checkout] ⚠ Low stock alert sent for variant: {item_name} (stok: {varian.stok})")
                    except Exception as e:
                        print(f"[Checkout] ⚠ Low stock notification failed: {e}")
            else:
                # Produk utama
                if produk.stok < quantity:
                    return jsonify({'success': False, 'message': f'Stok {produk.nama} tidak cukup'})
                
                # Kurangi stok produk utama
                produk.stok -= quantity
                item_name = produk.nama
                print(f"[Checkout] Item: {item_name} x{quantity}")
                
                # Check low stock for product
                if TELEGRAM_AVAILABLE and (not LICENSE_AVAILABLE or allows_telegram()):
                    try:
                        bot = get_telegram_bot()
                        threshold = app.config.get('TELEGRAM_NOTIFY_LOW_STOCK_THRESHOLD', 10)
                        if bot and produk.stok <= threshold:
                            bot.notify_low_stock_sync(
                                produk_nama=produk.nama,
                                stok=produk.stok,
                                kategori=produk.kategori.nama if produk.kategori else "Tanpa Kategori"
                            )
                            print(f"[Checkout] ⚠ Low stock alert sent for product: {item_name} (stok: {produk.stok})")
                    except Exception as e:
                        print(f"[Checkout] ⚠ Low stock notification failed: {e}")
            
            transaksi_item = TransaksiItem(
                transaksi_id=transaksi.id,
                produk_id=item['id'],
                jumlah=quantity,
                harga=item['price'],
                subtotal=item['price'] * quantity,
                varian_barcode=scanned_variant.get('barcode') if scanned_variant else None,
                varian_nama=scanned_variant.get('nama') if scanned_variant else None
            )
            db.session.add(transaksi_item)
        
        # Commit transaksi
        db.session.commit()

        if member:
            member.points = (member.points or 0) + points_earned
            member.total_spent = (member.total_spent or 0) + total
            db.session.commit()
        print("[Checkout] ✓ Transaction saved to database")
        
        # === TELEGRAM NOTIFICATION ===
        if TELEGRAM_AVAILABLE and (not LICENSE_AVAILABLE or allows_telegram()):
            try:
                bot = get_telegram_bot()
                if bot and app.config.get('TELEGRAM_NOTIFY_NEW_TRANSACTION', False):
                    bot.notify_new_transaction_sync(
                        kode_transaksi=kode_transaksi,
                        total=total,
                        payment_method=payment_method,
                        kasir=current_user.nama,
                        member_name=member.nama if member else member_manual
                    )
                    print("[Checkout] ✓ Telegram notification sent")
            except Exception as e:
                print(f"[Checkout] ⚠ Telegram notification failed: {e}")
        
        # === BACKUP SETELAH TRANSAKSI ===
        print("\n[Backup] Starting backup after transaction...")
        backup_success = backup_database()
        
        if backup_success:
            print("[Backup] ✓ Backup completed successfully!")
        else:
            print("[Backup] ✗ Backup failed!")
        
        print("="*50)
        print("CHECKOUT PROCESS COMPLETED")
        print("="*50 + "\n")
        
        return jsonify({
            'success': True,
            'kode_transaksi': kode_transaksi,
            'kembalian': bayar - total,
            'transaksi_id': transaksi.id,
            'total': total,
            'subtotal': subtotal,
            'discount_percent': 0,
            'discount_amount': 0,
            'points_earned': points_earned,
            'bayar': bayar,
            'payment_method': payment_method,
            'tanggal': transaksi.tanggal.strftime('%Y-%m-%d %H:%M:%S'),
            'kasir': current_user.nama
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"[Checkout] ✗ Error: {e}")
        import traceback
        error_detail = traceback.format_exc()
        print(f"[Checkout TRACEBACK] {error_detail}")
        
        return jsonify({
            'success': False, 
            'message': 'Terjadi kesalahan sistem'
        })

@app.route('/transaksi')
@login_required
def list_transaksi():
    # Get filter parameters
    tanggal_mulai = request.args.get('tanggal_mulai', '')
    tanggal_selesai = request.args.get('tanggal_selesai', '')
    payment_method = request.args.get('payment_method', '')
    kode = request.args.get('kode', '')

    page, per_page = _get_pagination_args(default_per_page=50, max_per_page=200)
    
    query = Transaksi.query
    
    # Apply filters
    if tanggal_mulai:
        query = query.filter(db.func.date(Transaksi.tanggal) >= tanggal_mulai)
    if tanggal_selesai:
        query = query.filter(db.func.date(Transaksi.tanggal) <= tanggal_selesai)
    if payment_method:
        query = query.filter(Transaksi.payment_method == payment_method)
    if kode:
        query = query.filter(Transaksi.kode_transaksi.like(f'%{kode}%'))
    
    pagination = _paginate_query(query, page=page, per_page=per_page, order_by=Transaksi.tanggal.desc())
    transaksi_list = pagination['items']
    
    return render_template('transaksi/list.html', 
                         transaksi_list=transaksi_list,
                         tanggal_mulai=tanggal_mulai,
                         tanggal_selesai=tanggal_selesai,
                         payment_method=payment_method,
                         kode=kode,
                         timezone_name=get_local_timezone_name(),
                         pagination=pagination)

@app.route('/transaksi/struk/<int:id>')
@login_required
def view_struk(id):
    transaksi = Transaksi.query.get_or_404(id)
    settings = {
        'store_name': Pengaturan.get('store_name', 'TOKO SEMBAKO'),
        'store_address': Pengaturan.get('store_address', 'Jl. Contoh No. 123'),
        'store_phone': Pengaturan.get('store_phone', '021-12345678'),
        'receipt_footer': Pengaturan.get('receipt_footer', 'Terima kasih atas kunjungan Anda')
    }
    return render_template('transaksi/struk.html', 
                         transaksi=transaksi, 
                         settings=settings,
                         timezone_name=get_local_timezone_name())

@app.route('/laporan')
@login_required
def laporan():
    if current_user.role != 'admin':
        flash('Akses ditolak! Hanya admin yang bisa melihat laporan.', 'danger')
        return redirect(url_for('index'))
    
    tanggal_mulai = request.args.get('tanggal_mulai', date.today().strftime('%Y-%m-%d'))
    tanggal_selesai = request.args.get('tanggal_selesai', date.today().strftime('%Y-%m-%d'))
    
    transaksi_list = Transaksi.query.filter(
        db.func.date(Transaksi.tanggal) >= tanggal_mulai,
        db.func.date(Transaksi.tanggal) <= tanggal_selesai
    ).order_by(Transaksi.tanggal).all()
    
    total_penjualan = sum(t.total for t in transaksi_list)
    total_transaksi = len(transaksi_list)
    total_keuntungan = 0
    for t in transaksi_list:
        for item in t.items:
            harga_beli = item.produk.harga_beli if item.produk else 0
            total_keuntungan += (item.harga - harga_beli) * item.jumlah
    
    # Chart data: Sales per day
    from collections import defaultdict
    sales_by_date = defaultdict(float)
    for t in transaksi_list:
        date_str = t.tanggal.strftime('%Y-%m-%d')
        sales_by_date[date_str] += t.total
    
    # Chart data: Payment method distribution
    payment_count = defaultdict(int)
    for t in transaksi_list:
        payment_count[t.payment_method or 'tunai'] += 1
    
    # Chart data: Top 5 products
    from collections import Counter
    product_sales = Counter()
    for t in transaksi_list:
        for item in t.items:
            product_sales[item.produk.nama] += item.subtotal
    top_products = product_sales.most_common(5)

    # Top 25 Members - berdasarkan bulan berjalan (reset setiap bulan)
    from datetime import datetime
    current_date = get_local_now()
    first_day_of_month = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # Last day of current month
    if current_date.month == 12:
        last_day_of_month = current_date.replace(year=current_date.year + 1, month=1, day=1, hour=23, minute=59, second=59, microsecond=999999)
    else:
        last_day_of_month = current_date.replace(month=current_date.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    top_members_rows = (db.session.query(
        Member.id,
        Member.nama,
        db.func.sum(Transaksi.total).label('total_spent'),
        db.func.count(Transaksi.id).label('total_transaksi')
    )
        .join(Transaksi, Transaksi.member_id == Member.id)
        .filter(
            Transaksi.tanggal >= first_day_of_month,
            Transaksi.tanggal < last_day_of_month
        )
        .group_by(Member.id, Member.nama)
        .order_by(db.func.sum(Transaksi.total).desc())
        .limit(25)
        .all())

    top_members = [
        {
            'nama': row.nama,
            'total_spent': row.total_spent or 0,
            'total_transaksi': row.total_transaksi or 0
        }
        for row in top_members_rows
    ]
    
    return render_template('laporan/index.html',
                         transaksi_list=transaksi_list,
                         tanggal_mulai=tanggal_mulai,
                         tanggal_selesai=tanggal_selesai,
                         total_penjualan=total_penjualan,
                         total_keuntungan=total_keuntungan,
                         total_transaksi=total_transaksi,
                         sales_by_date=dict(sales_by_date),
                         payment_count=dict(payment_count),
                         top_products=top_products,
                         top_members=top_members,
                         timezone_name=get_local_timezone_name())

@app.route('/laporan/export-excel', methods=['POST'])
@csrf.exempt
@login_required
def export_laporan_excel():
    """Export laporan keuangan ke format Excel"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        tanggal_mulai = request.form.get('tanggal_mulai', date.today().strftime('%Y-%m-%d'))
        tanggal_selesai = request.form.get('tanggal_selesai', date.today().strftime('%Y-%m-%d'))
        mode = request.form.get('mode', 'hari')  # hari, bulan, atau tahun
        
        # Debug log
        print(f'[EXPORT] tanggal_mulai={tanggal_mulai}, tanggal_selesai={tanggal_selesai}, mode={mode}')
        
        if mode == 'hari':
            wb = generate_laporan_hari(tanggal_mulai, tanggal_selesai)
        elif mode == 'bulan':
            wb = generate_laporan_bulan(tanggal_mulai, tanggal_selesai)
        elif mode == 'tahun':
            wb = generate_laporan_tahun(tanggal_mulai, tanggal_selesai)
        else:
            print(f'[EXPORT ERROR] Invalid mode: {mode}')
            return jsonify({'error': f'Mode tidak valid: {mode}'}), 400
        
        # Save to BytesIO buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        filename = f"laporan_keuangan_{mode}_{date.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    
    except Exception as e:
        print(f'[EXPORT ERROR] {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/laporan/trigger-daily-report', methods=['POST'])
@csrf.exempt
@login_required
def trigger_daily_report():
    """Manually trigger daily report generation (for testing)"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        generate_daily_report()
        return jsonify({'success': True, 'message': 'Laporan harian berhasil dibuat'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/laporan/list-reports', methods=['GET'])
@csrf.exempt
@login_required
def list_daily_reports():
    """List all generated daily reports"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        import os
        reports_dir = os.path.join(app.instance_path, 'reports')
        
        if not os.path.exists(reports_dir):
            return jsonify({'reports': [], 'total': 0}), 200
        
        # List all reports sorted by newest first
        files = os.listdir(reports_dir)
        reports = []
        
        for filename in sorted(files, reverse=True):
            if filename.endswith('.xlsx') and filename.startswith('laporan_harian_'):
                filepath = os.path.join(reports_dir, filename)
                file_size = os.path.getsize(filepath)
                file_date = os.path.getmtime(filepath)
                
                from datetime import datetime as dt_class
                mod_time = dt_class.fromtimestamp(file_date).strftime('%Y-%m-%d %H:%M:%S')
                
                reports.append({
                    'filename': filename,
                    'size': file_size,
                    'size_kb': round(file_size / 1024, 2),
                    'modified': mod_time,
                    'date': filename.replace('laporan_harian_', '').replace('.xlsx', '')
                })
        
        return jsonify({'reports': reports, 'total': len(reports)}), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/laporan/download-report/<filename>', methods=['GET'])
@csrf.exempt
@login_required
def download_daily_report(filename):
    """Download a specific daily report"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        import os
        # Security: only allow laporan_harian_*.xlsx files
        if not filename.startswith('laporan_harian_') or not filename.endswith('.xlsx'):
            return jsonify({'error': 'Invalid filename'}), 400
        
        reports_dir = os.path.join(app.instance_path, 'reports')
        filepath = os.path.join(reports_dir, filename)
        
        # Security: verify path is within reports directory
        if not os.path.abspath(filepath).startswith(os.path.abspath(reports_dir)):
            return jsonify({'error': 'Invalid path'}), 400
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        from flask import send_file
        return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_laporan_hari(tanggal_mulai, tanggal_selesai):
    """Generate laporan per hari dengan detail produk"""
    from openpyxl.styles import PatternFill, Border, Side, Font as OpenpyxlFont
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Per Hari"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = OpenpyxlFont(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    # Title
    ws['A1'] = "LAPORAN KEUANGAN TOKO"
    ws['A1'].font = OpenpyxlFont(bold=True, size=14)
    ws.merge_cells('A1:G1')
    row = 2
    
    ws[f'A{row}'] = f"Periode: {tanggal_mulai} s/d {tanggal_selesai}"
    ws.merge_cells(f'A{row}:G{row}')
    row += 2
    
    # Get all transactions for date range
    transaksi_list = Transaksi.query.filter(
        db.func.date(Transaksi.tanggal) >= tanggal_mulai,
        db.func.date(Transaksi.tanggal) <= tanggal_selesai
    ).order_by(Transaksi.tanggal).all()
    
    # Group by date
    from collections import defaultdict
    transaksi_by_date = defaultdict(list)
    for t in transaksi_list:
        date_key = t.tanggal.strftime('%Y-%m-%d')
        transaksi_by_date[date_key].append(t)
    
    # Totals for all data
    grand_total_pembayaran = 0
    grand_total_hpp = 0
    grand_total_keuntungan = 0
    
    # Process each date
    for date_key in sorted(transaksi_by_date.keys()):
        transaksi_hari = transaksi_by_date[date_key]
        
        # Date header
        ws[f'A{row}'] = "TANGGAL:" if row > 4 else f"{date_key}"
        ws[f'B{row}'] = date_key
        ws[f'A{row}'].font = OpenpyxlFont(bold=True)
        ws[f'B{row}'].font = OpenpyxlFont(bold=True)
        row += 1
        
        # Column headers for this day
        headers = ['No', 'Nama Barang', 'Jumlah', 'Harga Barang', 'Total Pembayaran', 'HPP', 'Keuntungan']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Data for this day - aggregate same products
        no = 1
        daily_total_pembayaran = 0
        daily_total_hpp = 0
        daily_total_keuntungan = 0
        
        # Collect product data
        product_data = {}  # {nama_produk: {jumlah, harga_jual, hpp, pembayaran, keuntungan}}
        
        for transaksi in transaksi_hari:
            for item in transaksi.items:
                nama_produk = item.produk.nama if item.produk else "Unknown"
                jumlah = item.jumlah
                harga_jual = item.harga
                total_pembayaran = item.subtotal
                hpp = item.produk.harga_beli if item.produk else 0
                keuntungan = (harga_jual - hpp) * jumlah
                
                # Aggregate the data
                if nama_produk not in product_data:
                    product_data[nama_produk] = {
                        'jumlah': 0,
                        'harga_jual': harga_jual,
                        'hpp': hpp,
                        'pembayaran': 0,
                        'keuntungan': 0
                    }
                
                product_data[nama_produk]['jumlah'] += jumlah
                product_data[nama_produk]['pembayaran'] += total_pembayaran
                product_data[nama_produk]['keuntungan'] += keuntungan
        
        # Write aggregated data to Excel
        for nama_produk, data in product_data.items():
            daily_total_pembayaran += data['pembayaran']
            daily_total_hpp += data['hpp'] * data['jumlah']
            daily_total_keuntungan += data['keuntungan']
            
            grand_total_pembayaran += data['pembayaran']
            grand_total_hpp += data['hpp'] * data['jumlah']
            grand_total_keuntungan += data['keuntungan']
            
            # Write data row
            ws[f'A{row}'] = no
            ws[f'B{row}'] = nama_produk
            ws[f'C{row}'] = data['jumlah']
            ws[f'D{row}'] = data['harga_jual']
            ws[f'E{row}'] = data['pembayaran']
            ws[f'F{row}'] = data['hpp']
            ws[f'G{row}'] = data['keuntungan']
            
            # Format currency columns
            for col in ['D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].border = border
            
            no += 1
            row += 1
        
        # Subtotal for this day
        subtotal_row = row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = OpenpyxlFont(bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'E{row}'] = daily_total_pembayaran
        ws[f'F{row}'] = daily_total_hpp
        ws[f'G{row}'] = daily_total_keuntungan
        
        for col in ['E', 'F', 'G']:
            cell = ws[f'{col}{row}']
            cell.font = OpenpyxlFont(bold=True)
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 2
    
    # Grand total at bottom
    ws[f'A{row}'] = "GRAND TOTAL"
    ws[f'A{row}'].font = OpenpyxlFont(bold=True, size=12)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'E{row}'] = grand_total_pembayaran
    ws[f'F{row}'] = grand_total_hpp
    ws[f'G{row}'] = grand_total_keuntungan
    
    for col in ['E', 'F', 'G']:
        cell = ws[f'{col}{row}']
        cell.font = OpenpyxlFont(bold=True, color="FFFFFF", size=12)
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    
    return wb

def generate_laporan_bulan(tanggal_mulai, tanggal_selesai):
    """Generate laporan per bulan (blok harian dengan summary bulanan)"""
    from openpyxl.styles import PatternFill, Border, Side, Font as OpenpyxlFont
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Per Bulan"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = OpenpyxlFont(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    # Title
    ws['A1'] = "LAPORAN KEUANGAN BULANAN"
    ws['A1'].font = OpenpyxlFont(bold=True, size=14)
    ws.merge_cells('A1:G1')
    row = 2
    
    ws[f'A{row}'] = f"Periode: {tanggal_mulai} s/d {tanggal_selesai}"
    ws.merge_cells(f'A{row}:G{row}')
    row += 2
    
    # Get all transactions
    transaksi_list = Transaksi.query.filter(
        db.func.date(Transaksi.tanggal) >= tanggal_mulai,
        db.func.date(Transaksi.tanggal) <= tanggal_selesai
    ).order_by(Transaksi.tanggal).all()
    
    from collections import defaultdict
    transaksi_by_date = defaultdict(list)
    for t in transaksi_list:
        date_key = t.tanggal.strftime('%Y-%m-%d')
        transaksi_by_date[date_key].append(t)
    
    month_totals = defaultdict(lambda: {'pembayaran': 0, 'hpp': 0, 'keuntungan': 0})
    
    # Process each date
    for date_key in sorted(transaksi_by_date.keys()):
        transaksi_hari = transaksi_by_date[date_key]
        month_key = date_key[:7]  # YYYY-MM
        
        # Date header
        ws[f'A{row}'] = f"TANGGAL: {date_key}"
        ws[f'A{row}'].font = OpenpyxlFont(bold=True, color="FFFFFF")
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        row += 1
        
        # Column headers
        headers = ['No', 'Nama Barang', 'Jumlah', 'Harga Barang', 'Total Pembayaran', 'HPP', 'Keuntungan']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Data - aggregate same products per day
        no = 1
        daily_pembayaran = 0
        daily_hpp = 0
        daily_keuntungan = 0
        
        # Collect product data for this day
        product_data = {}
        
        for transaksi in transaksi_hari:
            for item in transaksi.items:
                nama_produk = item.produk.nama if item.produk else "Unknown"
                jumlah = item.jumlah
                harga_jual = item.harga
                total_pembayaran = item.subtotal
                hpp = item.produk.harga_beli if item.produk else 0
                keuntungan = (harga_jual - hpp) * jumlah
                
                if nama_produk not in product_data:
                    product_data[nama_produk] = {
                        'jumlah': 0,
                        'harga_jual': harga_jual,
                        'hpp': hpp,
                        'pembayaran': 0,
                        'keuntungan': 0
                    }
                
                product_data[nama_produk]['jumlah'] += jumlah
                product_data[nama_produk]['pembayaran'] += total_pembayaran
                product_data[nama_produk]['keuntungan'] += keuntungan
        
        # Write aggregated data and accumulate totals
        for nama_produk, data in product_data.items():
            daily_pembayaran += data['pembayaran']
            daily_hpp += data['hpp'] * data['jumlah']
            daily_keuntungan += data['keuntungan']
            
            # Add to month total
            month_totals[month_key]['pembayaran'] += data['pembayaran']
            month_totals[month_key]['hpp'] += data['hpp'] * data['jumlah']
            month_totals[month_key]['keuntungan'] += data['keuntungan']
            
            # Write row
            ws[f'A{row}'] = no
            ws[f'B{row}'] = nama_produk
            ws[f'C{row}'] = data['jumlah']
            ws[f'D{row}'] = data['harga_jual']
            ws[f'E{row}'] = data['pembayaran']
            ws[f'F{row}'] = data['hpp']
            ws[f'G{row}'] = data['keuntungan']
            
            for col in ['D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].border = border
            
            no += 1
            row += 1
        
        # Daily subtotal
        ws[f'A{row}'] = "TOTAL HARI"
        ws[f'A{row}'].font = OpenpyxlFont(bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'E{row}'] = daily_pembayaran
        ws[f'F{row}'] = daily_hpp
        ws[f'G{row}'] = daily_keuntungan
        
        for col in ['E', 'F', 'G']:
            cell = ws[f'{col}{row}']
            cell.font = OpenpyxlFont(bold=True)
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 2
    
    # Monthly summary
    row += 1
    ws[f'A{row}'] = "RINGKASAN BULANAN"
    ws[f'A{row}'].font = OpenpyxlFont(bold=True, size=12, color="FFFFFF")
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    row += 1
    
    # Month header
    headers = ['Bulan', 'Total Pembayaran', 'Total HPP', 'Total Keuntungan']
    month_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = month_header_fill
        cell.font = OpenpyxlFont(bold=True, color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Month data
    for month_key in sorted(month_totals.keys()):
        ws[f'A{row}'] = month_key
        ws[f'B{row}'] = month_totals[month_key]['pembayaran']
        ws[f'C{row}'] = month_totals[month_key]['hpp']
        ws[f'D{row}'] = month_totals[month_key]['keuntungan']
        
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '#,##0.00'
            ws[f'{col}{row}'].border = border
        ws[f'A{row}'].border = border
        
        row += 1
    
    # Grand total for month report
    ws[f'A{row}'] = "TOTAL KESELURUHAN"
    ws[f'A{row}'].font = OpenpyxlFont(bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:A{row}')
    
    total_all_pembayaran = sum(v['pembayaran'] for v in month_totals.values())
    total_all_hpp = sum(v['hpp'] for v in month_totals.values())
    total_all_keuntungan = sum(v['keuntungan'] for v in month_totals.values())
    
    ws[f'B{row}'] = total_all_pembayaran
    ws[f'C{row}'] = total_all_hpp
    ws[f'D{row}'] = total_all_keuntungan
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = OpenpyxlFont(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.border = border
        cell.number_format = '#,##0.00'
    
    return wb

def generate_laporan_tahun(tanggal_mulai, tanggal_selesai):
    """Generate laporan per tahun dengan summary bulanan"""
    from openpyxl.styles import PatternFill, Border, Side, Font as OpenpyxlFont
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Per Tahun"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = OpenpyxlFont(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    # Title
    ws['A1'] = "LAPORAN KEUANGAN TAHUNAN"
    ws['A1'].font = OpenpyxlFont(bold=True, size=14)
    ws.merge_cells('A1:D1')
    row = 2
    
    ws[f'A{row}'] = f"Periode: {tanggal_mulai} s/d {tanggal_selesai}"
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # Get all transactions
    transaksi_list = Transaksi.query.filter(
        db.func.date(Transaksi.tanggal) >= tanggal_mulai,
        db.func.date(Transaksi.tanggal) <= tanggal_selesai
    ).order_by(Transaksi.tanggal).all()
    
    from collections import defaultdict
    month_totals = defaultdict(lambda: {'pembayaran': 0, 'hpp': 0, 'keuntungan': 0})
    
    # Calculate monthly totals
    for transaksi in transaksi_list:
        month_key = transaksi.tanggal.strftime('%Y-%m')
        for item in transaksi.items:
            hpp = item.produk.harga_beli if item.produk else 0
            keuntungan = (item.harga - hpp) * item.jumlah
            
            month_totals[month_key]['pembayaran'] += item.subtotal
            month_totals[month_key]['hpp'] += hpp * item.jumlah
            month_totals[month_key]['keuntungan'] += keuntungan
    
    # Table header
    headers = ['Bulan', 'Total Pembayaran', 'Total HPP', 'Total Keuntungan']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Month data
    for month_key in sorted(month_totals.keys()):
        ws[f'A{row}'] = month_key
        ws[f'B{row}'] = month_totals[month_key]['pembayaran']
        ws[f'C{row}'] = month_totals[month_key]['hpp']
        ws[f'D{row}'] = month_totals[month_key]['keuntungan']
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.border = border
            if col in ['B', 'C', 'D']:
                cell.number_format = '#,##0.00'
        
        row += 1
    
    # Grand total
    ws[f'A{row}'] = "TOTAL TAHUN"
    ws[f'A{row}'].font = OpenpyxlFont(bold=True, color="FFFFFF")
    
    total_all_pembayaran = sum(v['pembayaran'] for v in month_totals.values())
    total_all_hpp = sum(v['hpp'] for v in month_totals.values())
    total_all_keuntungan = sum(v['keuntungan'] for v in month_totals.values())
    
    ws[f'B{row}'] = total_all_pembayaran
    ws[f'C{row}'] = total_all_hpp
    ws[f'D{row}'] = total_all_keuntungan
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = OpenpyxlFont(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.border = border
        cell.number_format = '#,##0.00'
    
    return wb

# ==================== PENGATURAN ROUTES ====================

@app.route('/pengaturan', methods=['GET', 'POST'])
@login_required
def pengaturan():
    if current_user.role != 'admin':
        flash('Akses ditolak! Hanya admin yang bisa mengakses pengaturan.', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Update settings
        Pengaturan.set('store_name', request.form.get('store_name', ''))
        Pengaturan.set('store_address', request.form.get('store_address', ''))
        Pengaturan.set('store_phone', request.form.get('store_phone', ''))
        Pengaturan.set('receipt_footer', request.form.get('receipt_footer', ''))
        Pengaturan.set('tax_enabled', request.form.get('tax_enabled', 'false'))
        Pengaturan.set('tax_percentage', request.form.get('tax_percentage', '0'))
        
        flash('Pengaturan berhasil disimpan!', 'success')
        return redirect(url_for('pengaturan'))
    
    # Get current settings
    settings = {
        'store_name': Pengaturan.get('store_name', 'TOKO SEMBAKO'),
        'store_address': Pengaturan.get('store_address', 'Jl. Contoh No. 123'),
        'store_phone': Pengaturan.get('store_phone', '021-12345678'),
        'receipt_footer': Pengaturan.get('receipt_footer', 'Terima kasih atas kunjungan Anda'),
        'tax_enabled': Pengaturan.get('tax_enabled', 'false'),
        'tax_percentage': Pengaturan.get('tax_percentage', '10')
    }
    
    return render_template('pengaturan/index.html', settings=settings)

# ==================== BACKUP ROUTES ====================

@app.route('/admin/backup-now')
@login_required
def backup_now():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    if backup_database():
        flash('Backup berhasil dibuat!', 'success')
    else:
        flash('Gagal membuat backup!', 'danger')
    
    return redirect(url_for('index'))

@app.route('/admin/restore-backup')
@login_required
def restore_backup():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    # Implementasi restore nanti
    flash('Fitur restore dalam pengembangan', 'info')
    return redirect(url_for('index'))

# ==================== KAS ROUTES ====================



# ==================== ERROR HANDLERS ====================

@app.errorhandler(400)
def bad_request(e):
    return render_template('errors/400.html'), 400

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('errors/500.html'), 500

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    print(f"[CSRF] {e.description}")
    return render_template('errors/400.html'), 400

# ==================== INIT DATABASE ====================

def init_database():
    with app.app_context():
        db.create_all()

        # Lightweight migration for existing databases
        try:
            result = db.session.execute(text("PRAGMA table_info(transaksi)"))
            tx_columns = {row[1] for row in result}
            if 'member_id' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN member_id INTEGER"))
            if 'subtotal' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN subtotal REAL DEFAULT 0"))
            if 'discount_percent' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN discount_percent REAL DEFAULT 0"))
            if 'discount_amount' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN discount_amount REAL DEFAULT 0"))
            if 'points_earned' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN points_earned INTEGER DEFAULT 0"))

            result = db.session.execute(text("PRAGMA table_info(member)"))
            member_columns = {row[1] for row in result}
            if 'points' not in member_columns:
                db.session.execute(text("ALTER TABLE member ADD COLUMN points INTEGER DEFAULT 0"))
            if 'total_spent' not in member_columns:
                db.session.execute(text("ALTER TABLE member ADD COLUMN total_spent REAL DEFAULT 0"))

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[DB] Warning: gagal migrasi kolom member_id: {e}")

        # Indexes for large datasets (SQLite)
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_produk_kode ON produk(kode)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_produk_nama ON produk(nama)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_produk_kategori ON produk(kategori_id)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_produk_stok ON produk(stok)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_varian_barcode ON varian_produk(barcode_varian)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_varian_produk_id ON varian_produk(produk_id)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_member_nama ON member(nama)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_member_no_telp ON member(no_telp)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_transaksi_tanggal ON transaksi(tanggal)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_transaksi_kode ON transaksi(kode_transaksi)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_transaksi_payment ON transaksi(payment_method)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_transaksi_member_id ON transaksi(member_id)"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[DB] Warning: gagal membuat index: {e}")
        
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', nama='Administrator', role='admin')
            admin.set_password('Admin123')
            db.session.add(admin)
            
            kasir = User(username='kasir', nama='Kasir Toko', role='kasir')
            kasir.set_password('Kasir123')
            db.session.add(kasir)
            
            # Kategori
            kategori_list = [
                Kategori(nama='Sembako', deskripsi='Bahan pokok sehari-hari'),
                Kategori(nama='Minuman', deskripsi='Minuman berbagai jenis'),
                Kategori(nama='Snack', deskripsi='Makanan ringan'),
                Kategori(nama='Bahan Masak', deskripsi='Bahan-bahan untuk memasak'),
                Kategori(nama='Lainnya', deskripsi='Produk lainnya'),
            ]
            for kategori in kategori_list:
                if not Kategori.query.filter_by(nama=kategori.nama).first():
                    db.session.add(kategori)
            
            # Produk sample
            if not Produk.query.first():
                seed_names_only = (os.environ.get('SEED_PRODUCTS_NAMES_ONLY', 'false') or 'false').lower() == 'true'
                produk_sample = [
                    ('BRG001', 'Beras Premium 5kg', 1, 45000, 50000, 50),
                    ('BRG002', 'Minyak Goreng 2L', 4, 25000, 28000, 30),
                    ('BRG003', 'Gula Pasir 1kg', 1, 12000, 14000, 40),
                    ('BRG004', 'Aqua Gelas 240ml', 2, 500, 1000, 100),
                    ('BRG005', 'Indomie Goreng', 3, 2500, 3000, 60),
                ]
                
                for kode, nama, kategori_id, harga_beli, harga_jual, stok in produk_sample:
                    produk = Produk(
                        kode=kode,
                        nama=nama,
                        kategori_id=kategori_id,
                        harga_beli=0 if seed_names_only else harga_beli,
                        harga_jual=0 if seed_names_only else harga_jual,
                        stok=0 if seed_names_only else stok,
                        minimal_stok=5,
                        satuan='pcs'
                    )
                    db.session.add(produk)
            
            db.session.commit()
            print('✓ Database initialized')
            print('  Admin: admin / Admin123')
            print('  Kasir: kasir / Kasir123')
            print('  Note: Password harus mengandung huruf besar dan angka')
            
            # Backup pertama
            print('\n[Backup] Creating first backup...')
            backup_database()

# ==================== RUN APP ====================

if __name__ == '__main__':
    init_database()
    
    # Initialize Telegram Bot (only in main process, not in reloader)
    # Check if we're in the reloaded child process
    telegram_allowed_by_license = True
    if LICENSE_AVAILABLE:
        try:
            telegram_allowed_by_license = allows_telegram()
        except Exception:
            telegram_allowed_by_license = False

    if TELEGRAM_AVAILABLE and telegram_allowed_by_license and os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        bot_token = os.environ.get('TELEGRAM_BOT_TOKEN', '')
        admin_chat_ids_str = os.environ.get('TELEGRAM_ADMIN_CHAT_IDS', '')
        
        if bot_token and admin_chat_ids_str:
            admin_chat_ids = [id.strip() for id in admin_chat_ids_str.split(',') if id.strip()]
            if admin_chat_ids:
                print('\n' + '='*60)
                print('🤖 TELEGRAM BOT CONFIGURATION')
                print('='*60)
                print(f'✅ Bot Token: {bot_token[:10]}...')
                print(f'✅ Admin Chat IDs: {len(admin_chat_ids)} configured')
                
                try:
                    telegram_bot = initialize_telegram_bot(bot_token, admin_chat_ids, app)
                    if telegram_bot:
                        # Start bot in background
                        telegram_bot.start_bot_async()
                        print('✅ Telegram Bot started successfully!')
                        print('💡 Kirim /start ke bot untuk mulai monitoring')
                    else:
                        print('[ERROR] Failed to initialize Telegram Bot')
                except Exception as e:
                    print(f'[ERROR] Telegram Bot error: {e}')
                print('='*60)
            else:
                print('\n[WARNING] Telegram Bot: No admin chat IDs configured')
        else:
            print('\n[WARNING] Telegram Bot: Not configured (set TELEGRAM_BOT_TOKEN and TELEGRAM_ADMIN_CHAT_IDS)')
    elif TELEGRAM_AVAILABLE and not telegram_allowed_by_license:
        print('\n[WARNING] Telegram Bot: Disabled by license tier')
    elif TELEGRAM_AVAILABLE and not os.environ.get('WERKZEUG_RUN_MAIN'):
        print('\n[WARNING] Telegram Bot: Not configured (set TELEGRAM_BOT_TOKEN and TELEGRAM_ADMIN_CHAT_IDS)')
    elif not TELEGRAM_AVAILABLE:
        print('\n[WARNING] Telegram Bot: Not available (install: pip install python-telegram-bot)')
    
    print('\n' + '='*60)
    print('SERVER READY!')
    print('URL Lokal: http://localhost:5000')
    print('URL Jaringan: http://0.0.0.0:5000')
    print('Backup folder: backups/')
    print('='*60 + '\n')
    
    # Ensure database schema is up to date
    print('Checking database schema...')
    ensure_db_columns()
    print()
    
    # Start scheduler for daily saldo archive
    start_scheduler()
    
    app.run(host='0.0.0.0', debug=True, port=5000)